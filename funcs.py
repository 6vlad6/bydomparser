import os
import time

import openpyxl

import json

from selenium import webdriver
from selenium.webdriver.common.by import By


class BydomParser:

    def __init__(self, url):
        self.url = url

    def create_driver(self):
        """
        Функция создает драйвер для парсинга
        :param url: адрес страницы
        :return: драйвер
        """

        options = webdriver.ChromeOptions()
        options.add_argument("start-maximized")
        options.add_argument("--disable-blink-features=AutomationControlled")

        driver = webdriver.Chrome(options=options)
        driver.maximize_window()

        driver.get(self.url)

        return driver

    def create_folder(self):
        unixtime = int(time.time())  # время создания папки в unix-формате

        os.makedirs(str(unixtime))

        return unixtime

    def load_to_file(self, folder, cat_name, columns, data):

        workbook = openpyxl.Workbook()

        worksheet = workbook.active

        for i, header in enumerate(columns):
            worksheet.cell(row=1, column=i + 1, value=header)

        for row_num, row_data in enumerate(data):
            for col_num, cell_data in enumerate(row_data):
                try:
                    worksheet.cell(row=row_num + 2, column=col_num + 1, value=cell_data)
                except:
                    worksheet.cell(row=row_num + 2, column=col_num + 1, value=",".join(cell_data))

        file_name = f"{folder}/{cat_name}.xls"
        workbook.save(file_name)

        workbook.close()
